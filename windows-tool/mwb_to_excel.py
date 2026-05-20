"""
MWB (生活と奉仕の集会) ZIP → 週ごとタブの Excel 進行表 変換ツール

使い方:
    GUI:  python mwb_to_excel.py
    CLI:  python mwb_to_excel.py input.zip output.xlsx

入力: 生活と奉仕の集会ワークブック (JW Library エクスポートの ZIP)
      ZIP 内に mwb_J_YYYYMM_NN.txt 形式のファイルが含まれていること
出力: 各週がタブで分かれた Excel ファイル (.xlsx)
"""
import re
import sys
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl が見つかりません。pip install openpyxl を実行してください。")
    sys.exit(1)


# ── データ構造 ──────────────────────────
@dataclass
class RawItem:
    section: str
    number: str
    title: str
    minutes: str


@dataclass
class Week:
    file_id: str            # 例: 202605_02
    date_range: str = ''    # 例: 5月14日〜20日
    bible_chapter: str = '' # 例: 創世記X章
    opening_song: str = ''
    middle_song: str = ''
    closing_song: str = ''
    raw_items: list = field(default_factory=list)


# ── 正規化 ──────────────────────────────
ZEN_TO_HAN = str.maketrans('０１２３４５６７８９', '0123456789')

def normalize(s: str) -> str:
    """全角数字 → 半角数字"""
    return s.translate(ZEN_TO_HAN)


# ── パース処理（assignment.js の awParseWeekLines を移植） ──
RE_DATE      = re.compile(r'\d+月\d+')
RE_NUM_ITEM  = re.compile(r'^(\d+)\.\s+(.+?)（(\d+)分）')
RE_OPEN_SONG = re.compile(r'^(\d+)番の歌と祈り')
RE_MID_SONG  = re.compile(r'^(\d+)番の歌$')
RE_MINUTES   = re.compile(r'（(\d+)分）')
RE_NUM_PREFIX = re.compile(r'^(\d+)\.')

# 挿入ブロック判定用のキーワード（パースから除外する）
INSERT_BLOCK_START = ('挿入聖句', '読む聖句', '囲み')
INSERT_BLOCK_END = '終わり'


def parse_week_lines(lines: list) -> Optional[Week]:
    """テキスト行のリストから 1 週分の Week を抽出"""
    week = Week(file_id='')

    # 日付行を探す
    date_line_idx = -1
    for i, line in enumerate(lines):
        s = normalize(line.strip())
        if RE_DATE.search(s) and '日' in s and len(s) < 30:
            week.date_range = s
            date_line_idx = i
            break
    if date_line_idx == -1:
        return None

    # 直後の数行で聖句章を探す
    for j in range(date_line_idx + 1, min(date_line_idx + 5, len(lines))):
        s = normalize(lines[j].strip())
        if s:
            week.bible_chapter = s
            break

    # 本文をスキャン
    body = lines[date_line_idx:]
    current_section = ''
    in_inserted = False

    for line in body:
        raw = line.strip()
        s = normalize(raw)

        # 挿入ブロックのスキップ
        is_block_start = any(k in raw for k in INSERT_BLOCK_START)
        is_block_end   = INSERT_BLOCK_END in raw
        if is_block_start and not is_block_end:
            in_inserted = True
            continue
        if is_block_end:
            in_inserted = False
            continue
        if in_inserted:
            continue

        # セクション
        if '神の言葉の宝' in s:
            current_section = '神の言葉の宝'
            continue
        if '野外奉仕に励む' in s or '伝道を楽しもう' in s:
            current_section = '野外奉仕に励む'
            continue
        if 'クリスチャンとして生活する' in s:
            current_section = 'クリスチャンとして生活する'
            continue

        # 歌
        m = RE_OPEN_SONG.match(s)
        if m:
            if not week.opening_song:
                week.opening_song = m.group(1)
            else:
                week.closing_song = m.group(1)
            continue
        m = RE_MID_SONG.match(s)
        if m:
            week.middle_song = m.group(1)
            continue

        # 番号付き項目
        m = RE_NUM_ITEM.match(s)
        if m:
            week.raw_items.append(RawItem(
                section=current_section or '開会',
                number=m.group(1),
                title=m.group(2).strip(),
                minutes=m.group(3),
            ))
            continue

        # 番号なし定型項目
        if '開会の言葉' in s and '（' in s:
            dm = RE_MINUTES.search(s)
            week.raw_items.append(RawItem(
                section='開会', number='', title='開会の言葉',
                minutes=dm.group(1) if dm else '1',
            ))
            continue
        if '会衆の聖書研究' in s and '（' in s:
            dm = RE_MINUTES.search(s)
            week.raw_items.append(RawItem(
                section='クリスチャンとして生活する', number='',
                title='会衆の聖書研究', minutes=dm.group(1) if dm else '30',
            ))
            continue
        if '閉会の言葉' in s and '（' in s:
            dm = RE_MINUTES.search(s)
            week.raw_items.append(RawItem(
                section='クリスチャンとして生活する', number='',
                title='閉会の言葉', minutes=dm.group(1) if dm else '3',
            ))
            continue
        if '会衆の必要' in s:
            dm = RE_MINUTES.search(s)
            nm = RE_NUM_PREFIX.match(s)
            week.raw_items.append(RawItem(
                section='クリスチャンとして生活する',
                number=nm.group(1) if nm else '',
                title='会衆の必要',
                minutes=dm.group(1) if dm else '',
            ))
            continue

    return week


# ── ZIP 読込 ───────────────────────────
RE_ZIP_ENTRY = re.compile(r'mwb_J_(\d+)_(\d{2})\.txt$')


def extract_weeks_from_zip(zip_path: Path) -> list:
    """ZIP から mwb_J_*.txt を取り出して Week のリストを返す"""
    weeks = []
    with zipfile.ZipFile(zip_path, 'r') as zf:
        entries = []
        for name in zf.namelist():
            m = RE_ZIP_ENTRY.search(name)
            if m and not name.endswith('/'):
                entries.append((name, m.group(1), m.group(2)))
        entries.sort()

        for name, issue_yyyymm, week_idx in entries:
            with zf.open(name) as f:
                text = f.read().decode('utf-8-sig', errors='replace')
            lines = text.splitlines()
            week = parse_week_lines(lines)
            if not week or not week.date_range:
                continue
            week.file_id = f'{issue_yyyymm}_{week_idx}'
            weeks.append(week)
    return weeks


# ── Excel 出力 ─────────────────────────
SECTION_COLORS = {
    '開会':                       '4472C4',
    '神の言葉の宝':                'F4B084',
    '野外奉仕に励む':              'FFD966',
    'クリスチャンとして生活する':  'A9D18E',
    '閉会':                       '4472C4',
}

THIN = Side(border_style='thin', color='888888')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet_name_for(week: Week, fallback: str) -> str:
    """週用のシート名を作成。Excel 制限: 31文字以内・特殊文字禁止"""
    name = week.date_range or fallback
    # Excel が禁止する文字を除去
    for ch in '[]:*?/\\':
        name = name.replace(ch, '_')
    return name[:31]


def write_excel(weeks: list, output_path: Path) -> None:
    wb = Workbook()
    # デフォルトシート削除
    default = wb.active
    wb.remove(default)

    used_names = set()

    for w in weeks:
        base_name = sheet_name_for(w, w.file_id)
        name = base_name
        suffix = 2
        while name in used_names:
            name = f'{base_name[:28]}_{suffix}'
            suffix += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        _write_week_to_sheet(ws, w)

    wb.save(output_path)


def _write_week_to_sheet(ws, w: Week) -> None:
    """1 つの Week をシートに描画"""
    # 列幅
    ws.column_dimensions['A'].width = 8   # 時間
    ws.column_dimensions['B'].width = 6   # 番号
    ws.column_dimensions['C'].width = 50  # 項目
    ws.column_dimensions['D'].width = 10  # 分

    row = 1
    # ヘッダー
    ws.cell(row=row, column=1, value=w.date_range).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    if w.bible_chapter:
        ws.cell(row=row, column=1, value=w.bible_chapter).font = Font(size=11, color='555555')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    row += 1  # 空行

    # 進行表ヘッダ行
    headers = ['時間', '番号', '項目', '分']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='333333')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    row += 1

    # 各セクション + 項目
    minutes_offset = 0  # 集会開始からの経過分
    prev_section = ''

    items = _build_items(w)
    for it in items:
        # セクションヘッダ行
        if it['section'] != prev_section and it['section'] != '開会':
            section_label = it['section']
            sc = ws.cell(row=row, column=1, value=f'▼ {section_label}')
            sc.font = Font(bold=True, color='FFFFFF', size=11)
            sc.fill = PatternFill('solid', fgColor=SECTION_COLORS.get(section_label, '666666'))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = BORDER
            # クリスチャンセクションの最初は 47 分から再開
            if section_label == 'クリスチャンとして生活する':
                minutes_offset = 47
            row += 1
            prev_section = section_label

        # 時間表示
        h = 19 + minutes_offset // 60
        m = minutes_offset % 60
        time_str = f'{h}:{m:02d}'

        ws.cell(row=row, column=1, value=time_str).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=it['number']).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=it['title'])
        ws.cell(row=row, column=4, value=it['minutes']).alignment = Alignment(horizontal='center')
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER

        # 経過分を加算
        if it['type'] == 'song':
            minutes_offset += 5
        else:
            try:
                minutes_offset += int(it['minutes'] or 0)
            except ValueError:
                pass

        row += 1


def _build_items(w: Week) -> list:
    """歌・項目を時系列に整列"""
    items = []
    # 開会の歌
    if w.opening_song:
        items.append({
            'type': 'song', 'section': '開会',
            'number': '', 'title': f'{w.opening_song}番の歌と祈り',
            'minutes': '5',
        })
    # 各項目（開会の言葉は歌に統合済みなのでスキップ）
    for it in w.raw_items:
        if it.title == '開会の言葉':
            continue
        items.append({
            'type': 'item', 'section': it.section,
            'number': it.number, 'title': it.title,
            'minutes': it.minutes,
        })
    # 中間の歌をクリスチャンセクション先頭に挿入
    if w.middle_song:
        insert_idx = next(
            (i for i, it in enumerate(items) if it['section'] == 'クリスチャンとして生活する'),
            len(items),
        )
        items.insert(insert_idx, {
            'type': 'song', 'section': 'クリスチャンとして生活する',
            'number': '', 'title': f'{w.middle_song}番の歌',
            'minutes': '5',
        })
    # 閉会の歌
    if w.closing_song:
        items.append({
            'type': 'song', 'section': 'クリスチャンとして生活する',
            'number': '', 'title': f'{w.closing_song}番の歌と祈り',
            'minutes': '5',
        })
    return items


# ── GUI ────────────────────────────────
def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title('MWB → 進行表 Excel 変換')
    root.geometry('540x280')

    state = {'zip_path': None, 'output_path': None}

    # スタイル
    root.configure(padx=20, pady=20)

    title_lbl = tk.Label(
        root, text='生活と奉仕の集会 進行表 変換ツール',
        font=('Yu Gothic UI', 13, 'bold'), pady=8,
    )
    title_lbl.pack()

    desc_lbl = tk.Label(
        root,
        text='ZIP（mwb_J_*.txt を含む）を選んで Excel に変換します。',
        font=('Yu Gothic UI', 9), fg='#555',
    )
    desc_lbl.pack(pady=(0, 12))

    # ZIP 選択
    zip_frame = tk.Frame(root)
    zip_frame.pack(fill='x', pady=4)
    tk.Label(zip_frame, text='入力 ZIP:', width=10, anchor='w').pack(side='left')
    zip_var = tk.StringVar(value='（未選択）')
    tk.Label(zip_frame, textvariable=zip_var, fg='#333', anchor='w').pack(side='left', fill='x', expand=True)

    def pick_zip():
        path = filedialog.askopenfilename(
            title='MWB の ZIP を選択',
            filetypes=[('ZIP ファイル', '*.zip'), ('すべて', '*.*')],
        )
        if path:
            state['zip_path'] = Path(path)
            zip_var.set(state['zip_path'].name)

    tk.Button(zip_frame, text='選択...', command=pick_zip).pack(side='right')

    # 状態表示
    status_var = tk.StringVar(value='準備完了')
    status_lbl = tk.Label(root, textvariable=status_var, fg='#777', pady=12, font=('Yu Gothic UI', 9))
    status_lbl.pack(fill='x', pady=(8, 0))

    progress = ttk.Progressbar(root, mode='determinate', maximum=100)
    progress.pack(fill='x', pady=4)

    # 実行
    def run_convert():
        if not state['zip_path']:
            messagebox.showwarning('未選択', '先に ZIP を選んでください')
            return
        default_name = state['zip_path'].stem + '_進行表.xlsx'
        out = filedialog.asksaveasfilename(
            title='出力 Excel の保存先',
            initialfile=default_name,
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')],
        )
        if not out:
            return
        out_path = Path(out)
        try:
            status_var.set('ZIP を読み込み中...')
            progress['value'] = 20
            root.update()
            weeks = extract_weeks_from_zip(state['zip_path'])
            if not weeks:
                messagebox.showerror('エラー', 'mwb_J_*.txt が見つかりませんでした')
                status_var.set('準備完了')
                progress['value'] = 0
                return
            status_var.set(f'{len(weeks)} 週を Excel に書き出し中...')
            progress['value'] = 60
            root.update()
            write_excel(weeks, out_path)
            progress['value'] = 100
            status_var.set(f'完了: {out_path.name}')
            messagebox.showinfo(
                '完了',
                f'{len(weeks)} 週分を出力しました\n\n{out_path}',
            )
        except Exception as e:
            messagebox.showerror('エラー', f'変換に失敗しました\n\n{type(e).__name__}: {e}')
            status_var.set('エラー')
            progress['value'] = 0

    run_btn = tk.Button(
        root, text='Excel に変換', command=run_convert,
        font=('Yu Gothic UI', 11, 'bold'),
        bg='#047CBC', fg='white', padx=20, pady=8,
        relief='flat', cursor='hand2',
    )
    run_btn.pack(pady=12)

    root.mainloop()


# ── CLI ────────────────────────────────
def run_cli(zip_path: Path, output_path: Path) -> int:
    weeks = extract_weeks_from_zip(zip_path)
    if not weeks:
        print('ERROR: mwb_J_*.txt が見つかりませんでした', file=sys.stderr)
        return 1
    write_excel(weeks, output_path)
    print(f'{len(weeks)} 週分を {output_path} に出力しました')
    return 0


def main():
    if len(sys.argv) == 1:
        run_gui()
    elif len(sys.argv) == 3:
        zip_path = Path(sys.argv[1])
        output_path = Path(sys.argv[2])
        if not zip_path.is_file():
            print(f'ERROR: ファイルがありません: {zip_path}', file=sys.stderr)
            sys.exit(1)
        sys.exit(run_cli(zip_path, output_path))
    else:
        print(__doc__)
        sys.exit(1)


if __name__ == '__main__':
    main()
